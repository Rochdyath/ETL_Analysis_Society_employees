from airflow.decorators import dag, task
from airflow.utils.dates import days_ago
from pyspark.sql import SparkSession
import pyspark.sql.functions as f
from openpyxl import Workbook, load_workbook

@task
def get_budget_by_dept():
    spark = SparkSession.builder.getOrCreate()
    
    society = spark.read.option("header", True).csv("./dest_data/society.csv")
    
    budget_by_dept = society.groupby('dept_name').agg({'sum_salary': 'sum'})
    budget_by_dept = budget_by_dept.withColumnRenamed('sum(sum_salary)', 'budget')
    total = budget_by_dept.agg(f.sum("budget")).collect()[0][0]
    budget_by_dept = budget_by_dept.withColumn('budget_percent', f.col('budget')/total*100).sort("budget_percent", ascending=False)
    budget_by_dept.show()
    
    budget_by_dept = budget_by_dept.toPandas().to_dict(orient='records')

    spark.stop()
    
    return budget_by_dept


@task
def get_salary_by_dept():
    spark = SparkSession.builder.getOrCreate()
    
    society = spark.read.option("header", True).csv("./dest_data/society.csv")
    
    salary_by_dept = society.groupby('dept_name').agg({'avg_salary': 'mean'})
    salary_by_dept = salary_by_dept.withColumnRenamed('avg(avg_salary)', 'mean_salary')
    salary_by_dept.show()
    
    salary_by_dept = salary_by_dept.toPandas().to_dict(orient='records')

    spark.stop()
    
    return salary_by_dept


@task
def get_gender_by_dept():
    spark = SparkSession.builder.getOrCreate()
    
    society = spark.read.option("header", True).csv("./dest_data/society.csv")
    
    gender_by_dept = society.groupby('dept_name', 'gender').count().sort('dept_name', 'gender')
    total = society.count()
    gender_by_dept = gender_by_dept.withColumn('gender_percent', f.col('count')/total*100)
    gender_by_dept.show()
    
    gender_by_dept = gender_by_dept.toPandas().to_dict(orient='records')

    spark.stop()
    
    return gender_by_dept


def save_as_xlsx(df, sheet_name, file_name):
    spark = SparkSession.builder.getOrCreate()
    
    data = df.collect()
    columns = df.columns

    try:
        wb = load_workbook(file_name)
    except FileNotFoundError:
        wb = Workbook()

    if "Sheet" in wb: wb.remove(wb["Sheet"])
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
    
    ws.append(columns)
    
    for row in data:
        ws.append(list(row))
    
    wb.save(file_name)

    spark.stop()

@task
def save_dept_analysis(budget_by_dept, salary_by_dept, gender_by_dept):
    spark = SparkSession.builder.getOrCreate()
    
    budget_by_dept = spark.createDataFrame(budget_by_dept)
    salary_by_dept = spark.createDataFrame(salary_by_dept)
    gender_by_dept = spark.createDataFrame(gender_by_dept)

    dept_analysis = budget_by_dept.join(salary_by_dept, 'dept_name').join(gender_by_dept, 'dept_name').sort('dept_name', 'gender')
    dept_analysis.show()
    save_as_xlsx(dept_analysis, "department_analysis", "./dest_data/society_analysis.xlsx")
    
    spark.stop()


@task
def save_gender_salary_repartition():
    spark = SparkSession.builder.getOrCreate()

    society = spark.read.option("header", True).csv("./dest_data/society.csv")

    gender_salary_repartition = society.groupby('gender').agg({'avg_salary': 'mean'})
    gender_salary_repartition.show()
    save_as_xlsx(gender_salary_repartition, "gender_salary_repartition", "./dest_data/society_analysis.xlsx")

    spark.stop()

default_args = {
    'owner': 'airflow',
    'start_date': days_ago(1),
}

@dag(dag_id='society_analysis', default_args=default_args, schedule_interval='@daily', catchup=False)
def society_analysis():
    budget_by_dept = get_budget_by_dept()
    salary_by_dept = get_salary_by_dept()
    gender_by_dept = get_gender_by_dept()
    save_gender_salary_repartition()
    save_dept_analysis(budget_by_dept, salary_by_dept, gender_by_dept)

analysis = society_analysis()